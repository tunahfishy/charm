import pprint
import pandas as pd
import openpyxl 

# parse inputs into something that will be transformed into embeddings

# Read data from file 'filename.csv' 
data = pd.read_csv("cohort2bios.csv") 

# for row in data: 
#     print(row)
data.to_dict('series')
dict = {}
for i in range(52):
    dict[data["Name"][i]] = {"bio":data["Bio (~100 words)"][i]}
print(dict)
# for key,v in data.items():
#     print(key)
#     print()

# first_row = [] # The row where we stock the name of the column
# for col in range(1,ncols):
#     first_row.append(ws.cell(1,col) )
# print(first_row)
# transform the workbook to a list of dictionnary
# data = {}

# # make each row into all_bios = {name0: {"bio": bio}, name1: {"bio": bio}}
# for row in ws.values:
#     bio = {}
#     for col in range(1,ncols):
#         print(first_row[col] + " : " + ws.cell(row,col))
#         bio[first_row[col]]=ws.cell(row,col)
#     # name: []
#     data[bio[first_row[1]]] = bio

# print(data)

names = {
"Anubhav Bhardwaj": "Java Full Stack Developer",
"mayur chanpa":"Tax Associate 2 at Vialto Partners",
"Jay Bhadury":"Production Engineer at Reliance Industries Limited",
"tinah":"Quality Assurance Analyst at Hayat Kimya Nigeria",
"boyd": "Experienced customer service officer and a skilled virtual assistant . i can help you set up your payment gateway on selar and also help with settling up of paypal account.",
"caine": "Sql|Python|PowerBi|Azure",
"surya": "Student at Arya Institute of Engineering & Technology, Jaipur-2023",
"adarsh": "Test Lead at TCS and Automation Trainer",
"hi": "MBA in Marketing from “Osmania University ” worked as a Customer Support Professional for 11 Months in Sitel. Ex Sitel.",
"micky": "Software Test Engineer Trainee at QSpiders - Software Testing Training Institute",
"woah": "web developer, Web designer"
}